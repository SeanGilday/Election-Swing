from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
import csv
import json
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class helloHandler(BaseHTTPRequestHandler):

    def do_GET(self):
        workbk = load_workbook(filename='county_Feb22.xlsx')
        ws = workbk.active

        excelStats = []

        last_column = len(list(ws.columns))
        last_row = len(list(ws.rows))
        keys = ws[1]

        HOST = "https://www.elections.ny.gov/2022BallotProposal.html"
        URLback = "/EnrollmentCounty.html"
        pageback = requests.get(HOST)

        soup = BeautifulSoup(pageback.content, "html.parser")

        results = soup.find_all("div", id="midContent")
        print(results, "------------------\n")

        for row in range(2, last_row + 1):
            temp = dict()
            excelStats.append(temp)
            for col in range(1, len(ws[row])):
                temp[keys[col].value] = cellValue = ws[row][col].value
                cellValue = ws[row][col].value

        results = str(results)

        data = {"title": "Test", "description": results, "rows": excelStats}
        package = json.dumps(data)

        self.send_response(200)
        self.send_header('content-type', 'text/html')
        self.end_headers()
        #Arg my list JSON
        self.wfile.write(package.encode())
        print(package)


def main():

    workbk = load_workbook(filename='county_Feb22.xlsx')
    ws = workbk.active

    excelStats = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    keys = ws[1]

    for row in range(2, last_row + 1):
        temp = dict()
        excelStats.append(temp)
        for col in range(1, len(ws[row])):
            temp[keys[col].value] = cellValue = ws[row][col].value
            cellValue = ws[row][col].value
    data = json.dumps(excelStats)
    print(data)

    PORT = 8080
    server = HTTPServer(('', PORT), helloHandler)
    print('Server running on port %s' % PORT)
    server.serve_forever()


if __name__ == "__main__":
    main()

